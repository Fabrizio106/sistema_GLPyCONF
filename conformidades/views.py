from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db import transaction
from django.contrib import messages
from .models import TramiteConformidad, CertificadoConformidad, SedeConformidad
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.utils import timezone
from xhtml2pdf import pisa
from datetime import timedelta
from collections import OrderedDict
import json
from datetime import date, timedelta
from django.http import JsonResponse
from .forms import CertificadoForm, SedeConformidadForm
from django.contrib.auth.models import User
from django.db.models import Count
import openpyxl
from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

CAMPO_LABELS = {
    'categoria': 'Categoría', 'marca': 'Marca', 'modelo': 'Modelo',
    'color': 'Color', 'numero_vin': 'N° VIN', 'numero_serie': 'N° Serie',
    'numero_motor': 'Motor', 'anio_fabricacion': 'Año Fabricación',
    'anio_modelo': 'Año Modelo', 'ejes': 'Ejes', 'ruedas': 'Ruedas',
    'peso_bruto': 'Peso Bruto', 'peso_neto': 'Peso Neto', 'carga_util': 'Carga Útil',
    'carroceria': 'Carrocería', 'potencia': 'Potencia', 'combustible': 'Combustible',
    'formula_rodante': 'Fórmula Rodante', 'cilindrada': 'Cilindrada',
    'asientos': 'Asientos', 'pasajeros': 'Pasajeros', 'longitud': 'Longitud',
    'altura': 'Altura', 'ancho': 'Ancho', 'cilindros': 'Cilindros',
}

TIPOS_LABELS = {
    'INCORPORACION': 'INCORPORACIÓN', 'ADECUACION': 'ADECUACIÓN',
    'RECTIFICACION': 'RECTIFICACIÓN', 'MODIFICACION': 'MODIFICACIÓN',
}


def generar_norma_conformidad(tipo, campo, valor_nuevo, valor_anterior):
    if tipo != 'ADECUACION':
        return ""
    campo_lower = campo.lower()
    val_nuevo_up = str(valor_nuevo).upper().strip()
    if campo_lower == 'carroceria':
        TIPO_2008 = ['PICK UP','BARANDA','MINIBUS','MICROBÚS','MICROBUS',
                    'OMNIBUS URBANO','OMNIBUS INTERURBANO','OMNIBUS PANORAMICO',
                    'REMOLCADOR','GRUA','PANEL','CARGOBUS',
                    'TRIMOTO PASAJEROS','TRIMOTO CARGA','TRIMOTO DE PASAJEROS']
        TIPO_2006 = ['STATION WAGON','HATCH BACK','HATCHBACK','SUV',
                    'COUPE','MULTIPROPÓSITO','MULTIPROPOSITO','SEDAN',
                    'PLATAFORMA','VOLQUETE','M1']
        if any(c in val_nuevo_up for c in [x.upper() for x in TIPO_2008]):
            return "(Adecuación acorde a R.D. N°10476-2008-MTC/15)"
        elif any(c in val_nuevo_up for c in [x.upper() for x in TIPO_2006]):
            return "(Adecuación acorde a R.D. N°4848-2006-MTC/15)"
        else:
            return "(Adecuación a la directiva RD N° 04848-2006 MTC/15 y RD N° 10476-2008-MTC/15)"
    if campo_lower == 'combustible':
        return "(Adecuación a la RD N° 04848-2006 MTC/15)"
    if campo_lower == 'categoria':
        return "(Adecuación de acuerdo a la RD N° 04848-2006 MTC/15)"
    return ""


def generar_nota_conformidad(tipo, campo, valor_nuevo, valor_anterior):
    campo_lower = campo.lower()
    val_nuevo_up = str(valor_nuevo).upper().strip()
    val_ant_up = str(valor_anterior).upper().strip()

    if campo_lower == 'carroceria' and tipo == 'ADECUACION':
        TIPO_2008 = ['PICK UP','BARANDA','MINIBUS','MICROBÚS','MICROBUS',
                     'OMNIBUS URBANO','OMNIBUS INTERURBANO','OMNIBUS PANORAMICO',
                     'REMOLCADOR','GRUA','PANEL','CARGOBUS',
                     'TRIMOTO PASAJEROS','TRIMOTO CARGA','TRIMOTO DE PASAJEROS']
        TIPO_2006 = ['STATION WAGON','HATCH BACK','HATCHBACK','SUV',
                     'COUPE','MULTIPROPÓSITO','MULTIPROPOSITO','SEDAN',
                     'PLATAFORMA','VOLQUETE','M1']
        if any(c in val_nuevo_up for c in [x.upper() for x in TIPO_2008 + TIPO_2006]):
            return "EL VEHÍCULO INSPECCIONADO NO HA SUFRIDO MODIFICACIÓN ALGUNA EN SU ESTRUCTURA FÍSICA DEL CHASIS NI CARROCERÍA."

    if campo_lower == 'combustible':
        if 'GASOLINA' in val_nuevo_up:
            if 'GLP' in val_ant_up:
                return "EL SISTEMA BI-COMBUSTIBLE GLP HA SIDO DESMONTADO (KIT DE GLP)"
            elif 'GNV' in val_ant_up:
                return "EL VEHICULO YA NO CUENTA CON EL SISTEMA BI-COMBUSTIBLE GNV."
        if 'GLP' in val_nuevo_up and 'GLP' not in val_ant_up:
            return "EL VEHICULO CUENTA CON SISTEMA BI-COMBUSTIBLE GLP"
        if 'GNV' in val_nuevo_up and 'GNV' not in val_ant_up:
            return "EL VEHICULO MANTIENE EL SISTEMA BI-COMBUSTIBLE GNV"
        if tipo == 'RECTIFICACION':
            return "RECTIFICAR COMBUSTIBLE de acuerdo a la RD N 4848-2006-MTC/15."
    elif campo_lower == 'asientos':
        return "EL VEHÍCULO NO HA SUFRIDO MODIFICACIÓN ALGUNA CON RESPECTO AL NUMERO DE ASIENTOS."
    elif campo_lower == 'numero_motor':
        return ("EL NUEVO MOTOR ENSAMBLADO TIENE EL MISMO MODELO DEL ANTERIOR MOTOR "
                "POR LO CUAL, EL COMBUSTIBLE, CILINDRADA Y CILINDROS NO SE VEN AFECTADOS POR ESTE CAMBIO.")
    elif campo_lower == 'ruedas':
        if 'EXTRA' in val_ant_up or 'BALON' in val_ant_up:
            return "EL VEHICULO HA MODIFICADO DE RUEDAS TIPO EXTRA ANCHAS A RUEDAS TIPO CONVENCIONALES."
        else:
            return "EL VEHICULO HA MODIFICADO DE RUEDAS CONVENCIONALES A TIPO EXTRA ANCHAS (LLANTAS BALON)."
    elif campo_lower == 'categoria':
        if tipo == 'RECTIFICACION':
            return "RECTIFICAR CATEGORIA de acuerdo a la RD N 4848-2006-MTC/15."
    return ""


def render_to_pdf_conformidad(template_src, context_dict, filename="certificado"):
    template = get_template(template_src)
    html = template.render(context_dict)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}_CMD.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar PDF <pre>' + html + '</pre>')
    return response


def obtener_sedes_json():
    sedes_data = {
        str(s.id): {'anual_dias': s.fecha_anual, 'inicial_dias': s.fecha_inicial}
        for s in SedeConformidad.objects.all()
    }
    return json.dumps(sedes_data)


def formatear_valor_con_unidad(campo, valor):
    if not valor or str(valor).strip() == '':
        return str(valor)
    campo_lower = campo.lower()
    if campo_lower in ['longitud', 'altura', 'ancho']:
        return f"{str(valor).rstrip('0').rstrip('.')}m"
    if campo_lower in ['peso_neto', 'peso_bruto', 'carga_util']:
        try:
            return f"{int(float(str(valor)))}kg"
        except:
            return str(valor)
    return str(valor)


@login_required
def crear_certificado_conformidad(request):
    if request.method == 'POST':
        form = CertificadoForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    certificado = form.save(commit=False)
                    certificado.usuario = request.user
                    certificado.save()

                    # Capturamos todas las notas del formulario (vienen en el orden de los bloques visibles)
                    notas_bloques = request.POST.getlist('nota_modificacion_tramite[]')
                    
                    # Capturamos todos los tipos de trámites en orden
                    tipos_tramites_generales = request.POST.getlist('tipo_tramite[]')

                    mapa_tramites = request.POST.getlist('mapa_tramites[]')
                    
                    for item in mapa_tramites:
                        tipo, campo = item.split('|')
                        valor_nuevo = request.POST.get(f'valor_{tipo}_{campo}')
                        
                        if valor_nuevo:
                            nota_para_este_tramite = None
                            
                            # Si es una MODIFICACIÓN y requiere nota obligatoria
                            if tipo == 'MODIFICACION' and campo in ['combustible', 'numero_motor']:
                                try:
                                    # Encontramos en qué posición de los bloques generales se encuentra esta MODIFICACION
                                    # Para saber qué caja de texto (textarea) le corresponde.
                                    indices_modificacion = [i for i, x in enumerate(tipos_tramites_generales) if x == 'MODIFICACION']
                                    
                                    if indices_modificacion:
                                        # Tomamos el índice del primer bloque de modificación que coincida
                                        bloque_index = indices_modificacion[0] 
                                        
                                        if bloque_index < len(notas_bloques):
                                            nota_input = notas_bloques[bloque_index].strip()
                                            nota_para_este_tramite = nota_input if nota_input else None
                                except Exception as index_err:
                                    nota_para_este_tramite = None

                            # Guardamos el trámite enlazado al certificado con su respectiva nota
                            TramiteConformidad.objects.create(
                                certificado=certificado,
                                tipo_nombre=tipo,
                                campo_modificado=campo,
                                valor_nuevo=valor_nuevo,
                                nota=nota_para_este_tramite 
                            )

                    messages.success(request, "¡Certificado y trámites registrados con éxito!", extra_tags='conformidad')
                    return redirect('crear_certificado_conformidad')
            except Exception as e:
                form.add_error(None, f"Error al guardar: {e}")
    else:
        form = CertificadoForm()

    return render(request, 'conformidades/conformidad_form.html', {
        'form': form,
        'sedes_json': obtener_sedes_json()
    })


def consultar_ultima_conformidad(request):
    placa = request.GET.get('placa', None)
    if not placa:
        return JsonResponse({'error': 'No se proporcionó placa'}, status=400)
    ultima = CertificadoConformidad.objects.filter(placa=placa).order_by('-fecha_emision').first()
    if ultima:
        fecha_expiracion = ultima.fecha_emision + timedelta(days=365)
        dias_faltantes = (fecha_expiracion - date.today()).days
        return JsonResponse({
            'encontrado': True, 'placa': ultima.placa,
            'fecha_ultima': ultima.fecha_emision.strftime('%d/%m/%Y'),
            'dias_restantes': dias_faltantes if dias_faltantes > 0 else 0,
            'estado': 'VIGENTE' if dias_faltantes > 0 else 'EXPIRADO'
        })
    return JsonResponse({'encontrado': False})


@login_required
def historial_conformidades(request):
    certificados = (CertificadoConformidad.objects
                    .filter(usuario=request.user)
                    .order_by('-id')
                    .prefetch_related('tramites'))
    query = request.GET.get('q')
    sede_id = request.GET.get('sede')
    fecha_unica = request.GET.get('fecha')
    f_inicio = request.GET.get('inicio')
    f_fin = request.GET.get('fin')
    tipo_tramite = request.GET.get('tipo_tramite')
    if query:
        certificados = certificados.filter(placa__icontains=query)
    if sede_id:
        certificados = certificados.filter(sede_id=sede_id)
    if fecha_unica:
        certificados = certificados.filter(fecha_emision=fecha_unica)
    elif f_inicio and f_fin:
        certificados = certificados.filter(fecha_emision__range=[f_inicio, f_fin])
    if tipo_tramite:
        certificados = certificados.filter(tramites__tipo_nombre=tipo_tramite).distinct()
    return render(request, 'conformidades/historial_conformidades.html', {
        'certificados': certificados,
        'sedes': SedeConformidad.objects.all(),
        'campo_labels': CAMPO_LABELS,
    })


@login_required
def eliminar_conformidad(request, pk):
    certificado = get_object_or_404(CertificadoConformidad, pk=pk)
    if request.method == 'POST':
        certificado.delete()
        messages.success(request, "Certificado eliminado correctamente.")
    return redirect('historial_conformidades')


@login_required
def descargar_pdf_conformidad(request, pk):
    certificado = get_object_or_404(CertificadoConformidad, pk=pk)
    tramites = list(certificado.tramites.all())

    bloques_dict = OrderedDict()
    for t in tramites:
        bloques_dict.setdefault(t.tipo_nombre, []).append(t)

    bloques = []
    notas = []

    for tipo, items in bloques_dict.items():
        primer = items[0]
        primer_campo = primer.campo_modificado
        primer_val_new = primer.valor_nuevo
        valor_anterior = formatear_valor_con_unidad(
            primer_campo, str(getattr(certificado, primer_campo, '') or '')
        )

        adicionales = []
        for item in items[1:]:
            norma_ad = generar_norma_conformidad(
                item.tipo_nombre, item.campo_modificado, item.valor_nuevo,
                str(getattr(certificado, item.campo_modificado, '') or '')
            )
            adicionales.append({
                'label': CAMPO_LABELS.get(item.campo_modificado, item.campo_modificado.upper()),
                'valor_nuevo': formatear_valor_con_unidad(item.campo_modificado, item.valor_nuevo),
                'norma': norma_ad,
            })

        if (primer_campo.lower() == 'combustible'
                and 'GASOLINA' in str(primer_val_new).upper()
                and ('GLP' in valor_anterior.upper() or 'GNV' in valor_anterior.upper())):
            if certificado.peso_neto:
                adicionales.append({'label': 'PESO NETO', 'valor_nuevo': f'{float(certificado.peso_neto) - 30:.0f}kg', 'norma': ''})
            if certificado.carga_util:
                adicionales.append({'label': 'CARGA UTIL', 'valor_nuevo': f'{float(certificado.carga_util) + 30:.0f}kg', 'norma': ''})

        norma_primer = generar_norma_conformidad(tipo, primer_campo, primer_val_new, valor_anterior)

        for item in items:
            valor_ant_item = str(getattr(certificado, item.campo_modificado, '') or '')
            
            # Si es MODIFICACION de combustible o motor Y tiene nota manual, solo usa la manual
            if (item.tipo_nombre == 'MODIFICACION' 
                    and item.campo_modificado in ['combustible', 'numero_motor']
                    and item.nota and item.nota.strip()):
                if item.nota.strip() not in notas:
                    notas.append(item.nota.strip())
            else:
                # Para todo lo demás, usa la nota automática normal
                nota_auto = generar_nota_conformidad(item.tipo_nombre, item.campo_modificado, item.valor_nuevo, valor_ant_item)
                if nota_auto and nota_auto not in notas:
                    notas.append(nota_auto)

        bloques.append({
            'tipo': TIPOS_LABELS.get(tipo, tipo),
            'tipo_raw': tipo,
            'primer_campo': primer_campo,
            'primer_label': CAMPO_LABELS.get(primer_campo, primer_campo.upper()),
            'valor_anterior': valor_anterior,
            'primer_val_new': formatear_valor_con_unidad(primer_campo, primer.valor_nuevo),
            'norma_primer': norma_primer,
            'adicionales': adicionales,
        })

    fecha_pdf = certificado.fecha_emision or timezone.localdate()
    fecha_str = f"{fecha_pdf.day:02d}-{fecha_pdf.month:02d}-{fecha_pdf.year}"

    categoria_final = certificado.categoria or ''
    for t in tramites:
        if t.campo_modificado.lower() == 'categoria':
            categoria_final = t.valor_nuevo
            break

    tipo_transporte = 'PERSONAS' if categoria_final.upper().startswith('M') else 'MERCANCÍAS'

    return render_to_pdf_conformidad(
        'conformidades/pdf_conformidad.html',
        {'certificado': certificado, 'bloques': bloques, 'notas': notas, 'fecha_str': fecha_str, 'tipo_transporte': tipo_transporte},
        filename=certificado.placa or 'certificado'
    )


@login_required
def editar_conformidad(request, pk):
    certificado = get_object_or_404(CertificadoConformidad, pk=pk)
    if request.method == 'POST':
        form = CertificadoForm(request.POST, instance=certificado)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Eliminamos los trámites anteriores para recrearlos con los nuevos cambios
                    certificado.tramites.all().delete()
                    editado = form.save(commit=False)
                    editado.usuario = request.user
                    # editado.fecha_emision = timezone.localdate()
                    editado.save()

                    mapa_tramites = request. POST. getlist( 'mapa_tramites[]')
                    notas_bloques = request. POST. getlist('nota_modificacion_tramite[]')
                    tipos_generales = request.POST. getlist('tipo_tramite[]')
                    vistos = set()

                    for item in mapa_tramites:
                        try:
                            tipo, campo = item.split('|')
                            clave = f'{tipo}|{campo}'
                            if clave in vistos:
                                continue
                            vistos.add(clave)
                            valor_nuevo = request.POST.get(f'valor_{tipo}_{campo}')
                            if valor_nuevo and valor_nuevo.strip():
                                nota = None
                                if tipo == 'MODIFICACION' and campo in ['combustible', 'numero_motor']:
                                    try:
                                        indices = [i for i, x in enumerate(tipos_generales) if x == 'MODIFICACION']
                                        if indices and indices[0] < len(notas_bloques):
                                            nota_txt = notas_bloques[indices[0]].strip()
                                            nota = nota_txt if nota_txt else None
                                    except:
                                        nota = None

                            TramiteConformidad.objects.create(
                                certificado=editado,
                                tipo_nombre=tipo,
                                campo_modificado=campo,
                                valor_nuevo=valor_nuevo.strip(),
                                nota=nota
                            )
                        except ValueError:
                            continue

                    messages.success(request, "Certificado actualizado correctamente.", extra_tags='conformidad')
                    origen = request.POST.get('origen')
                    if origen == 'reporte':
                        return redirect('reporte_cmd_admin')
                    else:
                        return redirect('historial_conformidades')
            except Exception as e:
                form.add_error(None, f"Error al guardar: {e}")
    else:
        form = CertificadoForm(instance=certificado)

    import json as _json
    # CORRECCIÓN AQUÍ: Agregamos 'nota' a la extracción de datos para el JS
    tramites_existentes = list(
        certificado.tramites.values('tipo_nombre', 'campo_modificado', 'valor_nuevo', 'nota')
    )
    return render(request, 'conformidades/conformidad_form.html', {
        'form': form, 'editando': True,
        'sedes_json': obtener_sedes_json(),
        'tramites_json': _json.dumps(tramites_existentes),
    })

def obtener_sedes_json():
    sedes_data = {
        str(s.id): s.regla_fecha
        for s in SedeConformidad.objects.all()
    }
    return json.dumps(sedes_data)


@login_required
def gestion_sedes_conformidad(request):
    sedes = SedeConformidad.objects.all()
    if request.method == 'POST':
        form = SedeConformidadForm(request.POST)
        if form.is_valid():
            form.save()
            # AGREGAMOS extra_tags='sede_info'
            messages.success(request, "Sede de conformidad registrada con éxito.", extra_tags='sede_info')
            return redirect('gestion_sedes_conformidad')
    else:
        form = SedeConformidadForm()
        
    return render(request, 'conformidades/sede_conformidad_form.html', {
        'form': form,
        'sedes': sedes,
        'editando': False
    })

@login_required
def editar_sede_conformidad(request, pk):
    sede = get_object_or_404(SedeConformidad, pk=pk)
    sedes = SedeConformidad.objects.all()
    if request.method == 'POST':
        form = SedeConformidadForm(request.POST, instance=sede)
        if form.is_valid():
            form.save()
            # AGREGAMOS extra_tags='sede_info'
            messages.success(request, "Sede de conformidad actualizada correctamente.", extra_tags='sede_info')
            return redirect('gestion_sedes_conformidad')
    else:
        form = SedeConformidadForm(instance=sede)
        
    return render(request, 'conformidades/sede_conformidad_form.html', {
        'form': form,
        'sedes': sedes,
        'editando': sede
    })

@login_required
def eliminar_sede_conformidad(request, pk):
    sede = get_object_or_404(SedeConformidad, pk=pk)
    if request.method == 'POST':
        sede.delete()
        messages.success(request, "Sede eliminada correctamente.", extra_tags='sede_info')
    return redirect('gestion_sedes_conformidad')

@login_required
def reporte_cmd_admin(request):
    query = request.GET.get('q')
    usuario_id = request.GET.get('usuario')
    sede_id = request.GET.get('sede')
    fecha_unica = request.GET.get('fecha')
    f_inicio = request.GET.get('inicio')
    f_fin = request.GET.get('fin')
    tipo_tramite = request.GET.get('tipo_tramite')

    filtros_activos = any([query, usuario_id, sede_id, fecha_unica, f_inicio, f_fin, tipo_tramite])

    if filtros_activos:
        certificados = CertificadoConformidad.objects.all().order_by('-id').prefetch_related('tramites', 'sede', 'usuario')

        if query:
            certificados = certificados.filter(placa__icontains=query)
        if usuario_id:
            certificados = certificados.filter(usuario_id=usuario_id)
        if sede_id:
            certificados = certificados.filter(sede_id=sede_id)
        if fecha_unica:
            certificados = certificados.filter(fecha_emision=fecha_unica)
        elif f_inicio and f_fin:
            certificados = certificados.filter(fecha_emision__range=[f_inicio, f_fin])
        if tipo_tramite:
            certificados = certificados.filter(tramites__tipo_nombre=tipo_tramite).distinct()

        total_encontrados = certificados.count()
        total_modificaciones = certificados.filter(tramites__tipo_nombre='MODIFICACION').distinct().count()
        total_adecuaciones = certificados.filter(tramites__tipo_nombre='ADECUACION').distinct().count()
        total_incorporaciones = certificados.filter(tramites__tipo_nombre='INCORPORACION').distinct().count()
        total_rectificaciones = certificados.filter(tramites__tipo_nombre='RECTIFICACION').distinct().count()

    else:
        certificados = CertificadoConformidad.objects.none()
        total_encontrados = 0
        total_modificaciones = 0
        total_adecuaciones = 0
        total_incorporaciones = 0
        total_rectificaciones = 0

    return render(request, 'conformidades/reporte_cmd.html', {
        'certificados': certificados,
        'usuarios': User.objects.all(),
        'sedes': SedeConformidad.objects.all(),
        'total_encontrados': total_encontrados,
        'total_modificaciones': total_modificaciones,
        'total_adecuaciones': total_adecuaciones,
        'total_incorporaciones': total_incorporaciones,
        'total_rectificaciones': total_rectificaciones,
    })

@login_required
def descargar_excel_cmd(request):
    # 1. Filtros
    query = request.GET.get('q')
    usuario_id = request.GET.get('usuario')
    sede_id = request.GET.get('sede')
    fecha_unica = request.GET.get('fecha')
    f_inicio = request.GET.get('inicio')
    f_fin = request.GET.get('fin')
    tipo_tramite = request.GET.get('tipo_tramite')

    filtros_activos = any([query, usuario_id, sede_id, fecha_unica, f_inicio, f_fin, tipo_tramite])

    if filtros_activos:
        certificados = CertificadoConformidad.objects.all().order_by('-id').prefetch_related('tramites', 'sede', 'usuario')

        if query:
            certificados = certificados.filter(placa__icontains=query)
        if usuario_id:
            certificados = certificados.filter(usuario_id=usuario_id)
        if sede_id:
            certificados = certificados.filter(sede_id=sede_id)
        if fecha_unica:
            certificados = certificados.filter(fecha_emision=fecha_unica)
        elif f_inicio and f_fin:
            certificados = certificados.filter(fecha_emision__range=[f_inicio, f_fin])
        if tipo_tramite:
            certificados = certificados.filter(tramites__tipo_nombre=tipo_tramite).distinct()
    else:
        certificados = CertificadoConformidad.objects.none()

    # 2. Configuración de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte CMD"

    # ESTILOS: Fuente, Alineación, Fondo y ahora BORDES
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="FF7300", end_color="FF7300", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Creamos un borde delgado para los 4 lados de la celda
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [
        "Nº", "Sede", "Placa", "Certificador", 
        "Trámites Realizados", "Fecha Emisión", "Fecha Certificación", "Hora"
    ]
    ws.append(headers)

    # Aplicamos estilos y bordes a la CABECERA
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = alignment_center
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # 3. Volcado de datos con conversor de zona horaria
    for index, c in enumerate(certificados, start=1):
        lista_tramites = []
        for t in c.tramites.all():
            lista_tramites.append(f"[{t.tipo_nombre}] {t.campo_modificado.upper()}: {t.valor_nuevo}")
        tramites_str = " | ".join(lista_tramites) if lista_tramites else "Sin trámites"

        # Transformamos la hora UTC a la hora local exacta
        fecha_registro_local = timezone.localtime(c.fecha_registro) if c.fecha_registro else None

        fecha_emision_str = fecha_registro_local.strftime('%d/%m/%Y') if fecha_registro_local else "—"
        fecha_cert_str = c.fecha_emision.strftime('%d/%m/%Y') if c.fecha_emision else "—"
        hora_str = fecha_registro_local.strftime('%H:%M') if fecha_registro_local else "—"
        
        certificador_str = c.usuario.get_full_name() if c.usuario and c.usuario.get_full_name() else (c.usuario.username if c.usuario else "—")

        row = [
            index,
            c.sede.nombre_sede.upper() if c.sede else "—",
            c.placa.upper() if c.placa else "—",
            certificador_str,
            tramites_str,
            fecha_emision_str,
            fecha_cert_str,
            hora_str
        ]
        ws.append(row)
        
        # Aplicamos bordes a todas las celdas de la fila que acabamos de agregar
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reporte_conformidades_cmd.xlsx"'
    wb.save(response)
    return response