from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl
from .models import CertificadoGLP, SedeConfiguracion
from .forms import SedeConfiguracionForm, CertificadoGLPForm
from datetime import date, timedelta
from django.utils import timezone
import json
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User, Group
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.contrib.auth.decorators import login_required

def obtener_sedes_json():
    sedes_queryset = SedeConfiguracion.objects.all()
    sedes_data = {
        str(s.id): {
            'anual_dias': s.fecha_anual,
            'inicial_dias': s.fecha_inicial
        } for s in sedes_queryset
    }
    return json.dumps(sedes_data)

@login_required
def registrar_certificado_glp(request):
    if request.method == 'POST':
        form = CertificadoGLPForm(request.POST)
        if form.is_valid():
            certificado = form.save(commit=False)
            certificado.usuario = request.user
            certificado.save()
            messages.success(request, "¡Certificado registrado con éxito!")
            return redirect('registrar_glp') 
    else:
        form = CertificadoGLPForm()

    return render(request, 'glp/formulario_glp.html', {
        'form': form,
        'sedes_json': obtener_sedes_json()
    })


def render_to_pdf(template_src, context_dict={}, filename="documento"):
    template = get_template(template_src)
    html = template.render(context_dict) 
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Error <pre>' + html + '</pre>')
    return response

@login_required
def historial_glp(request):
    certificados = CertificadoGLP.objects.filter(usuario=request.user).order_by('-id')
    query = request.GET.get('q') # por placa
    sede_id = request.GET.get('sede')      # por sede
    fecha_unica = request.GET.get('fecha') # por dia
    f_inicio = request.GET.get('inicio')   # por rango
    f_fin = request.GET.get('fin')

    tiene_filtro = False

    if query:
        certificados = certificados.filter(placa__icontains=query)
        tiene_filtro = True
    if sede_id:
        certificados = certificados.filter(sede_id=sede_id)
        tiene_filtro = True
    if fecha_unica:
        certificados = certificados.filter(fecha_emision=fecha_unica)
        tiene_filtro = True
    elif f_inicio and f_fin:
        certificados = certificados.filter(fecha_emision__range=[f_inicio, f_fin])
        tiene_filtro = True

    if not tiene_filtro:
        certificados = certificados.filter(fecha_emision=timezone.localdate())
    
    sedes = SedeConfiguracion.objects.all()
        
    return render(request, 'glp/historial_glp.html', {
        'certificados': certificados,
        'sedes': sedes,
    })

@login_required
def descargar_pdf_glp(request, pk):
    certificado = get_object_or_404(CertificadoGLP, pk=pk)
    return render_to_pdf('glp/pdf_certificado.html', {'certificado': certificado}, filename=certificado.placa)


@login_required
@staff_member_required
def gestion_sedes(request, pk=None):
    sede_instancia = get_object_or_404(SedeConfiguracion, pk=pk) if pk else None
    
    if request.method == 'POST':
        form = SedeConfiguracionForm(request.POST, instance=sede_instancia)
        if form.is_valid():
            form.save()
            return redirect('gestion_sedes')
    else:
        form = SedeConfiguracionForm(instance=sede_instancia)

    sedes = SedeConfiguracion.objects.all()
    return render(request, 'glp/sede_configuracion.html', {
        'form': form,
        'sedes': sedes,
        'editando': sede_instancia
    })

@login_required
def eliminar_sede(request, pk):
    sede = get_object_or_404(SedeConfiguracion, pk=pk)
    sede.delete()
    return redirect('gestion_sedes')

@login_required
def editar_certificado_glp(request, pk):
    certificado = get_object_or_404(CertificadoGLP, pk=pk)
    if request.method == 'POST':
        form = CertificadoGLPForm(request.POST, instance=certificado) #con el instance 
        if form.is_valid():
            certificado_editado = form.save(commit=False)
            certificado_editado.usuario = request.user
            certificado_editado.save()
            messages.success(request, "Certificado actualizado correctamente.")
            return redirect('historial_glp')
    else:
        form = CertificadoGLPForm(instance=certificado)

    return render(request, 'glp/formulario_glp.html', {
        'form': form, 
        'editando': True,
        'sedes_json': obtener_sedes_json()
    })

@login_required
def eliminar_certificado_glp(request, pk):
    certificado = get_object_or_404(CertificadoGLP, pk=pk)
    if request.method == 'POST':
        certificado.delete()
        messages.success(request, "Certificado eliminado correctamente.")
    return redirect('historial_glp')

@login_required
@staff_member_required
def gestion_usuarios(request, user_id=None):
    usuario_editar = None
    if user_id:
        usuario_editar = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        rol_nombre = request.POST.get('rol')

        if usuario_editar:
            usuario_editar.username = username
            usuario_editar.first_name = first_name
            usuario_editar.last_name = last_name
            if password: # Solo actualiza contra si se escribe algo
                usuario_editar.password = make_password(password)
            usuario_editar.save()

            usuario_editar.groups.clear() 
            grupo = Group.objects.get_or_create(name=rol_nombre)[0]
            usuario_editar.groups.add(grupo)
            messages.success(request, f"Usuario {username} actualizado.", extra_tags='usuario')
        else:
            if User.objects.filter(username=username).exists():
                messages.error(request, "El nombre de usuario ya existe.", extra_tags='usuario')
            else:
                nuevo_usuario = User.objects.create(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    password=make_password(password)
                )
                grupo = Group.objects.get_or_create(name=rol_nombre)[0]
                nuevo_usuario.groups.add(grupo)
                messages.success(request, "Usuario registrado con éxito.", extra_tags='usuario')
        
        return redirect('gestion_usuarios')

    usuarios = User.objects.all().order_by('-id')
    return render(request, 'gestion_usuarios.html', {
        'usuarios': usuarios,
        'usuario_editar': usuario_editar 
    })

@login_required
def eliminar_usuario(request, user_id):

    usuario = get_object_or_404(User, id=user_id)
    
    if usuario == request.user:
        messages.error(request, "No puedes eliminar tu propio usuario mientras estás en sesión.")
    else:
        nombre_usuario = usuario.username
        usuario.delete()
        messages.warning(request, f"Usuario eliminado.", extra_tags='usuario')
    
    return redirect('gestion_usuarios')

@login_required
@staff_member_required
def reporte_glp_admin(request):

    q = request.GET.get('q')
    user_id = request.GET.get('usuario')
    sede_id = request.GET.get('sede')
    f_unica = request.GET.get('fecha')
    f_inicio = request.GET.get('inicio')
    f_fin = request.GET.get('fin')

    hubo_busqueda = any([q, user_id, sede_id, f_unica, (f_inicio and f_fin)])

    if hubo_busqueda:
        certificados = CertificadoGLP.objects.all().order_by('-id')
        if q:
            certificados = certificados.filter(placa__icontains=q)
        if user_id:
            certificados = certificados.filter(usuario_id=user_id)
        if sede_id:
            certificados = certificados.filter(sede_id=sede_id)
        if f_unica:
            certificados = certificados.filter(fecha_emision=f_unica)
        elif f_inicio and f_fin:
            certificados = certificados.filter(fecha_emision__range=[f_inicio, f_fin])
    else:
        certificados = CertificadoGLP.objects.none()

    usuarios = User.objects.filter(is_staff=False)
    sedes = SedeConfiguracion.objects.all()

    total_encontrados = certificados.count()
    total_anuales = certificados.filter(tipo_certificado='ANUAL').count()
    total_iniciales = certificados.filter(tipo_certificado='INICIAL').count()

    return render(request, 'glp/reporte_general_glp.html', {
        'certificados': certificados,
        'usuarios': usuarios,
        'sedes': sedes, 
        'total_encontrados': total_encontrados,
        'total_anuales': total_anuales,
        'total_iniciales': total_iniciales
    })

def consultar_ultima_conformidad(request):
    placa = request.GET.get('placa', None)
    if not placa:
        return JsonResponse({'error': 'No se proporcionó placa'}, status=400)
    
    # Buscamos la última conformidad por fecha de emisión
    ultima = CertificadoGLP.objects.filter(placa=placa).order_by('-fecha_certificacion').first()
    
    if ultima:
        fecha_expiracion = ultima.fecha_emision + timedelta(days=365)
        dias_faltantes = (fecha_expiracion - date.today()).days
        
        return JsonResponse({
            'encontrado': True,
            'placa': ultima.placa,
            'fecha_ultima': ultima.fecha_emision.strftime('%d/%m/%Y'),
            'dias_restantes': dias_faltantes if dias_faltantes > 0 else 0,
            'estado': 'VIGENTE' if dias_faltantes > 0 else 'EXPIRADO'
        })
    else:
        return JsonResponse({'encontrado': False})

def descargar_excel_glp(request):
    certificados = CertificadoGLP.objects.all()

    # 1. Filtros (Igual que antes)
    q = request.GET.get('q')
    sede_id = request.GET.get('sede')
    fecha = request.GET.get('fecha')
    inicio = request.GET.get('inicio')
    fin = request.GET.get('fin')

    if q:
        certificados = certificados.filter(placa__icontains=q)
    if sede_id:
        certificados = certificados.filter(sede_id=sede_id)
    if fecha:
        certificados = certificados.filter(fecha_emision=fecha)
    if inicio and fin:
        certificados = certificados.filter(fecha_emision__range=[inicio, fin])

    # 2. Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte GLP"

    # ==========================================
    # DEFINICIÓN DE ESTILOS (El diseño)
    # ==========================================
    # Color de fondo azul (estilo corporativo) para el encabezado
    color_encabezado = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    fuente_encabezado = Font(color="FFFFFF", bold=True) # Letra blanca y negrita
    
    # Borde negro delgado para todas las celdas (formato cuadro)
    borde_fino = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Centrado de texto
    alineacion_centro = Alignment(horizontal="center", vertical="center")

    # ==========================================
    # CREAR LOS ENCABEZADOS
    # ==========================================
    encabezados = ['N°', 'Sede', 'Ciudad', 'Placa', 'Certificador', 'Tipo', 'Fecha Emisión', 'Fecha Certificación', 'Hora']
    ws.append(encabezados)

    # Aplicar diseño solo a la primera fila (los encabezados)
    for celda in ws[1]:
        celda.fill = color_encabezado
        celda.font = fuente_encabezado
        celda.alignment = alineacion_centro
        celda.border = borde_fino

    # ==========================================
    # LLENAR DATOS CON DISEÑO
    # ==========================================
    contador = 1 # Iniciamos el contador
    
    for c in certificados:
        nombre_certificador = str(getattr(c, 'usuario', ''))
        if c.fecha_registro:
            hora_exacta = timezone.localtime(c.fecha_registro).strftime('%H:%M:%S')
        else:
            hora_exacta = ''
        fila = [
            contador,
            c.sede.nombre_sede if c.sede else 'Sin Sede',
            c.ciudad_glp_pdf,
            c.placa,
            nombre_certificador,
            c.tipo_certificado,
            c.fecha_emision.strftime('%d/%m/%Y') if c.fecha_emision else '',
            c.fecha_certificacion.strftime('%d/%m/%Y') if c.fecha_certificacion else '',
            hora_exacta
        ]
        ws.append(fila)
        contador += 1 # Aumentamos el contador para la siguiente fila

    # Aplicar el borde y centrado a todas las filas de datos que acabamos de crear
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for celda in row:
            celda.border = borde_fino
            celda.alignment = alineacion_centro

    # (Opcional) Ajustar el ancho de las columnas automáticamente para que no se vea apretado
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Damos un poquito de respiro (+4) al ancho de la columna
        ws.column_dimensions[col_letter].width = max_length + 4

    # 3. Descargar
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Certificados_GLP.xlsx"'
    wb.save(response)
    
    return response